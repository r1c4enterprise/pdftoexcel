"""
Flask Web App — PDF to Excel Converter
Tampilan web untuk upload PDF dan download hasil Excel.
"""

from flask import Flask, request, jsonify, render_template, send_file
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import os

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024


# ─────────────────────────────────────────────────────────────────────────────
def try_parse_number(value: str):
    """Konversi string angka ke int/float. Format ID dan EN didukung."""
    if not isinstance(value, str):
        return value
    v = value.strip()
    if v in ("", "-"):
        return value

    negative = False
    if v.startswith("(") and v.endswith(")"):
        v = v[1:-1]
        negative = True
    elif v.startswith("-"):
        v = v[1:]
        negative = True

    v = re.sub(r"[Rp$€£¥\s]", "", v)

    if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)?$", v):      # 1.000.000,50
        v = v.replace(".", "").replace(",", ".")
    elif re.match(r"^\d{1,3}(,\d{3})*(\.\d+)?$", v):    # 1,000,000.50
        v = v.replace(",", "")
    elif re.match(r"^\d+(,\d+)$", v):                    # 1000,50
        v = v.replace(",", ".")
    elif not re.match(r"^\d+(\.\d+)?$", v):
        return value

    try:
        num = float(v)
        if negative:
            num = -num
        return int(num) if num == int(num) else num
    except ValueError:
        return value


def extract_tables(pdf_bytes: bytes) -> list:
    """Ekstrak semua tabel dari PDF, kembalikan list of dict."""
    results = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if not tables:
                continue
            for tbl_idx, table in enumerate(tables, start=1):
                if not table or len(table) < 2:
                    continue

                header = [str(h).strip() if h else "" for h in table[0]]
                rows   = [
                    [str(c).strip() if c else "" for c in row]
                    for row in table[1:]
                ]

                # Deduplikasi header
                seen = {}
                clean_header = []
                for col in header:
                    if col in seen:
                        seen[col] += 1
                        clean_header.append(f"{col}_{seen[col]}")
                    else:
                        seen[col] = 0
                        clean_header.append(col)

                # Buang baris kosong semua
                rows = [r for r in rows if any(c.strip() for c in r)]

                # Konversi angka
                parsed_rows = []
                for row in rows:
                    parsed_rows.append([try_parse_number(c) for c in row])

                results.append({
                    "page":        page_num,
                    "table_index": tbl_idx,
                    "headers":     clean_header,
                    "rows":        parsed_rows,
                })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    """Halaman utama — tampilan web upload."""
    return render_template("index.html")


@app.route("/convert-download", methods=["POST"])
def convert_download():
    """
    Terima PDF, proses, kembalikan file Excel langsung untuk didownload.
    """
    if "file" not in request.files:
        return jsonify({"error": "Tidak ada file yang dikirim."}), 400

    f = request.files["file"]
    if f.filename == "" or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Hanya file PDF yang diterima."}), 400

    pdf_bytes = f.read()
    original_name = os.path.splitext(f.filename)[0]

    try:
        tables = extract_tables(pdf_bytes)
        excel_bytes = build_excel(tables)
    except Exception as e:
        return jsonify({"error": f"Gagal memproses PDF: {str(e)}"}), 500

    return send_file(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{original_name}.xlsx",
    )


@app.route("/convert", methods=["POST"])
def convert():
    """Endpoint JSON untuk Google Apps Script."""
    if "file" not in request.files:
        return jsonify({"error": "Tidak ada file yang dikirim."}), 400

    f = request.files["file"]
    if f.filename == "" or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Hanya file PDF yang diterima."}), 400

    pdf_bytes = f.read()

    try:
        tables = extract_tables(pdf_bytes)
    except Exception as e:
        return jsonify({"error": f"Gagal memproses PDF: {str(e)}"}), 500

    return jsonify({
        "status":       "ok",
        "total_tables": len(tables),
        "tables":       tables,
    })


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "PDF API berjalan."})


# ─────────────────────────────────────────────────────────────────────────────
def build_excel(tables: list) -> bytes:
    """Buat file Excel dari list tabel, kembalikan bytes."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        current_row = 0
        for item in tables:
            page    = item["page"]
            tbl_idx = item["table_index"]
            headers = item["headers"]
            rows    = item["rows"]

            df = pd.DataFrame(rows, columns=headers)

            # Label pemisah
            label_df = pd.DataFrame([[f"[ Halaman {page} — Tabel {tbl_idx} ]"]])
            label_df.to_excel(writer, sheet_name="Laporan Keuangan",
                              startrow=current_row, index=False, header=False)
            current_row += 1

            df.to_excel(writer, sheet_name="Laporan Keuangan",
                        startrow=current_row, index=False, header=True)
            current_row += len(df) + 1 + 2

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Laporan Keuangan"]
    _apply_styles(ws)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.read()


def _apply_styles(ws):
    COLOR_LABEL  = "2F5496"
    COLOR_HEADER = "4472C4"
    COLOR_ODD    = "FFFFFF"
    COLOR_EVEN   = "D9E1F2"
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    label_rows  = set()
    header_rows = set()

    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and str(cell.value).startswith("[ Halaman"):
            label_rows.add(cell.row)
    for r in label_rows:
        header_rows.add(r + 1)

    data_row_counters = {}

    for row in ws.iter_rows():
        row_num   = row[0].row
        is_label  = row_num in label_rows
        is_header = row_num in header_rows

        if not is_label and not is_header:
            block_start = max((h for h in header_rows if h < row_num), default=0)
            data_row_counters.setdefault(block_start, 0)
            data_row_counters[block_start] += 1
            is_even = (data_row_counters[block_start] % 2 == 0)
        else:
            is_even = False

        for cell in row:
            if is_label:
                cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
                cell.fill      = PatternFill("solid", fgColor=COLOR_LABEL)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif is_header:
                cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border
            else:
                cell.font = Font(size=10, name="Calibri")
                cell.fill = PatternFill("solid", fgColor=COLOR_EVEN if is_even else COLOR_ODD)
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment    = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) and cell.value != int(cell.value) else '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    for row in ws.row_dimensions.values():
        row.height = 18


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
