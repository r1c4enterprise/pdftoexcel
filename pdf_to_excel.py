"""
PDF to Excel Converter
Konversi laporan keuangan PDF (banyak tabel) ke satu sheet Excel yang rapi.
"""

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers as xl_numbers
import re
import os
import sys


# ─── KONFIGURASI ──────────────────────────────────────────────────────────────

INPUT_PDF  = "input.pdf"       # ganti dengan nama file PDF kamu
OUTPUT_XLS = "output.xlsx"     # nama file Excel hasil

# ──────────────────────────────────────────────────────────────────────────────


def try_parse_number(value: str):
    """
    Coba konversi string ke angka (int atau float).
    Mendukung format:
      - Indonesia : 1.000.000,50  →  1000000.5
      - Inggris   : 1,000,000.50  →  1000000.5
      - Negatif   : (1.000)  atau  -1.000
    Kembalikan angka jika berhasil, atau string asli jika tidak bisa dikonversi.
    """
    if not isinstance(value, str):
        return value

    v = value.strip()

    if v == "" or v == "-":
        return value

    # Tangani angka negatif dalam kurung: (1.000) → -1000
    negative = False
    if v.startswith("(") and v.endswith(")"):
        v = v[1:-1]
        negative = True
    elif v.startswith("-"):
        v = v[1:]
        negative = True

    # Hapus simbol mata uang dan spasi
    v = re.sub(r"[Rp$€£¥\s]", "", v)

    # Deteksi format Indonesia (titik sebagai ribuan, koma sebagai desimal)
    # Contoh: 1.000.000,50
    if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)?$", v):
        v = v.replace(".", "").replace(",", ".")
    # Format Inggris (koma sebagai ribuan, titik sebagai desimal)
    # Contoh: 1,000,000.50
    elif re.match(r"^\d{1,3}(,\d{3})*(\.\d+)?$", v):
        v = v.replace(",", "")
    # Angka biasa dengan koma desimal saja: 1000,50
    elif re.match(r"^\d+(,\d+)$", v):
        v = v.replace(",", ".")
    # Kalau tidak cocok pola angka sama sekali, kembalikan asli
    elif not re.match(r"^\d+(\.\d+)?$", v):
        return value

    try:
        num = float(v)
        if negative:
            num = -num
        # Kembalikan int jika tidak ada desimal
        return int(num) if num == int(num) else num
    except ValueError:
        return value


def extract_tables_from_pdf(pdf_path: str) -> list[dict]:
    """
    Baca semua tabel dari setiap halaman PDF.
    Return list of dict: { 'page': int, 'table_index': int, 'df': DataFrame }
    """
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"📄 Total halaman: {total_pages}")

        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            if not tables:
                print(f"   Halaman {page_num}: tidak ada tabel, dilewati.")
                continue

            print(f"   Halaman {page_num}: ditemukan {len(tables)} tabel.")

            for tbl_idx, table in enumerate(tables, start=1):
                if not table or len(table) < 2:
                    # Abaikan tabel kosong atau hanya header
                    continue

                # Baris pertama sebagai header, sisanya data
                header = table[0]
                rows   = table[1:]

                # Bersihkan None → string kosong
                header = [str(h).strip() if h is not None else "" for h in header]
                rows   = [
                    [str(c).strip() if c is not None else "" for c in row]
                    for row in rows
                ]

                # Deduplikasi nama kolom kalau ada yang sama
                seen = {}
                clean_header = []
                for col in header:
                    if col in seen:
                        seen[col] += 1
                        clean_header.append(f"{col}_{seen[col]}")
                    else:
                        seen[col] = 0
                        clean_header.append(col)

                df = pd.DataFrame(rows, columns=clean_header)

                # Buang baris yang seluruhnya kosong
                df = df[~df.apply(lambda r: r.str.strip().eq("").all(), axis=1)]
                df.reset_index(drop=True, inplace=True)

                # Konversi kolom yang isinya angka ke tipe numerik
                for col in df.columns:
                    df[col] = df[col].apply(try_parse_number)

                results.append({
                    "page":        page_num,
                    "table_index": tbl_idx,
                    "df":          df,
                })

    return results


def write_to_excel(tables: list[dict], output_path: str) -> None:
    """
    Tulis semua tabel ke satu sheet Excel dengan pemisah antar tabel.
    """
    if not tables:
        print("⚠️  Tidak ada tabel yang ditemukan di PDF.")
        return

    # ── Tulis DataFrame ke Excel dulu pakai pandas ──
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        current_row = 0  # baris aktif di sheet (0-indexed untuk pandas)

        for item in tables:
            page      = item["page"]
            tbl_idx   = item["table_index"]
            df        = item["df"]

            # -- Label pemisah --
            label_df = pd.DataFrame(
                [[f"[ Halaman {page} — Tabel {tbl_idx} ]"]]
            )
            label_df.to_excel(
                writer,
                sheet_name="Laporan Keuangan",
                startrow=current_row,
                index=False,
                header=False,
            )
            current_row += 1  # satu baris label

            # -- Data tabel --
            df.to_excel(
                writer,
                sheet_name="Laporan Keuangan",
                startrow=current_row,
                index=False,
                header=True,
            )
            current_row += len(df) + 1 + 2  # header(1) + data + 2 baris kosong

    # ── Buka lagi untuk styling ──
    wb = load_workbook(output_path)
    ws = wb["Laporan Keuangan"]

    _apply_styles(ws)

    wb.save(output_path)
    print(f"\n✅ File Excel tersimpan: {output_path}")


def _apply_styles(ws) -> None:
    """Terapkan warna, font, border, dan lebar kolom ke worksheet."""

    # Warna
    COLOR_LABEL  = "2F5496"   # biru tua  → label pemisah
    COLOR_HEADER = "4472C4"   # biru      → header tabel
    COLOR_ODD    = "FFFFFF"   # putih     → baris ganjil
    COLOR_EVEN   = "D9E1F2"   # biru muda → baris genap

    # Border tipis
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Deteksi baris label vs header vs data
    # Label: sel A mengandung "[ Halaman"
    # Header: baris tepat setelah label
    # Data: sisanya

    label_rows  = set()
    header_rows = set()

    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and str(cell.value).startswith("[ Halaman"):
            label_rows.add(cell.row)

    for r in label_rows:
        header_rows.add(r + 1)

    data_row_counters = {}   # untuk warna selang-seling

    for row in ws.iter_rows():
        row_num = row[0].row
        is_label  = row_num in label_rows
        is_header = row_num in header_rows

        # Hitung urutan data per blok tabel
        if not is_label and not is_header:
            # Cari blok tabel terdekat di atasnya
            block_start = max((h for h in header_rows if h < row_num), default=0)
            data_row_counters.setdefault(block_start, 0)
            data_row_counters[block_start] += 1
            is_even = (data_row_counters[block_start] % 2 == 0)
        else:
            is_even = False

        for cell in row:
            # Lewati sel kosong total
            if cell.value is None and not is_label and not is_header:
                continue

            # Font
            if is_label:
                cell.font = Font(
                    bold=True, color="FFFFFF", size=11, name="Calibri"
                )
                cell.fill = PatternFill("solid", fgColor=COLOR_LABEL)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )
            elif is_header:
                cell.font = Font(
                    bold=True, color="FFFFFF", size=10, name="Calibri"
                )
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
            else:
                cell.font = Font(size=10, name="Calibri")
                fill_color = COLOR_EVEN if is_even else COLOR_ODD
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = border
                # Format angka: rata kanan + format ribuan
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center", wrap_text=False
                    )
                    # Tampilkan dengan pemisah ribuan, 2 desimal jika float
                    if isinstance(cell.value, float) and cell.value != int(cell.value):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

    # Auto-fit lebar kolom
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max_len + 4, 50)  # max 50 karakter
        ws.column_dimensions[col_letter].width = adjusted

    # Freeze baris pertama (opsional, karena multi-tabel freeze kurang relevan)
    # ws.freeze_panes = "A2"

    # Tinggi baris default
    for row in ws.row_dimensions.values():
        row.height = 18


def main():
    pdf_path = INPUT_PDF

    # Boleh juga pass argumen dari command line: python pdf_to_excel.py namafile.pdf
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]

    if not os.path.exists(pdf_path):
        print(f"❌ File tidak ditemukan: {pdf_path}")
        print("   Taruh file PDF di folder yang sama, lalu ubah INPUT_PDF di atas.")
        sys.exit(1)

    print(f"🔍 Membaca PDF: {pdf_path}")
    tables = extract_tables_from_pdf(pdf_path)
    print(f"\n📊 Total tabel ditemukan: {len(tables)}")

    write_to_excel(tables, OUTPUT_XLS)


if __name__ == "__main__":
    main()
