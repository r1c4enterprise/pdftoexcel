"""
PDF to Excel Converter — Streamlit App
Upload laporan keuangan PDF, download hasilnya sebagai Excel.
"""

import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import time

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PDF → Excel Converter",
    page_icon="⚡",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Inject CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

* { font-family: 'Inter', sans-serif; }

/* Hide streamlit default elements */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 2rem; max-width: 760px; }

/* Hero section */
.hero {
    text-align: center;
    padding: 2rem 1rem 2rem;
}
.hero-badge {
    display: inline-block;
    background: linear-gradient(135deg, #4F8BF9, #A855F7);
    color: white;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    padding: 6px 16px;
    border-radius: 99px;
    margin-bottom: 1.2rem;
}
.hero-title {
    font-size: 2.8rem;
    font-weight: 800;
    line-height: 1.15;
    margin: 0 0 1rem;
    color: #FAFAFA;
}
.gradient-text {
    background: linear-gradient(135deg, #4F8BF9, #A855F7);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-sub {
    color: #718096;
    font-size: 1rem;
    font-weight: 400;
    margin-bottom: 2rem;
    line-height: 1.7;
}

/* Features */
.features {
    display: flex;
    gap: 0.6rem;
    margin-bottom: 1.5rem;
    flex-wrap: wrap;
    justify-content: center;
}
.feature-chip {
    background: #1E2130;
    border: 1px solid #2D3748;
    border-radius: 99px;
    padding: 6px 14px;
    font-size: 0.78rem;
    color: #A0AEC0;
}

/* Glow divider */
.glow-line {
    height: 2px;
    background: linear-gradient(90deg, transparent, #4F8BF9, #A855F7, transparent);
    margin: 1.5rem 0;
    border-radius: 99px;
}

/* Upload label */
.upload-label {
    font-size: 0.9rem;
    font-weight: 600;
    color: #A0AEC0;
    margin-bottom: 0.5rem;
    display: flex;
    align-items: center;
    gap: 8px;
}

/* Stats row */
.stat-card {
    background: #1E2130;
    border: 1px solid #2D3748;
    border-radius: 14px;
    padding: 1.2rem;
    text-align: center;
}
.stat-icon { font-size: 1.4rem; margin-bottom: 0.3rem; }
.stat-value {
    font-size: 1.4rem;
    font-weight: 700;
    color: #4F8BF9;
    display: block;
}
.stat-label {
    font-size: 0.7rem;
    color: #718096;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* Result card */
.result-card {
    background: linear-gradient(135deg, #0D1B2A, #1A1F35);
    border: 1px solid #1E4D8C;
    border-radius: 16px;
    padding: 1.5rem;
    margin-top: 1rem;
}
.result-title {
    font-size: 1.1rem;
    font-weight: 700;
    color: #63B3ED;
    margin-bottom: 1rem;
}

/* Footer */
.footer {
    text-align: center;
    color: #4A5568;
    font-size: 0.78rem;
    padding: 2rem 0 0;
}

/* Override streamlit upload button */
[data-testid="stFileUploader"] > div {
    border: 2px dashed #2D3748 !important;
    border-radius: 14px !important;
    background: #1A1F2E !important;
    transition: border-color 0.3s;
}
[data-testid="stFileUploader"] > div:hover {
    border-color: #4F8BF9 !important;
    background: #1E2538 !important;
}

/* Primary button */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #4F8BF9, #7C3AED) !important;
    border: none !important;
    border-radius: 12px !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 20px rgba(79,139,249,0.35) !important;
    transition: transform 0.15s, box-shadow 0.15s !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 30px rgba(79,139,249,0.5) !important;
}

/* Download button */
[data-testid="stDownloadButton"] > button {
    width: 100% !important;
    background: linear-gradient(135deg, #059669, #0D9488) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 20px rgba(5,150,105,0.35) !important;
}
[data-testid="stDownloadButton"] > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 30px rgba(5,150,105,0.5) !important;
}

/* Progress bar */
.stProgress > div > div > div {
    background: linear-gradient(90deg, #4F8BF9, #A855F7) !important;
    border-radius: 99px !important;
}

/* Metric */
[data-testid="stMetric"] {
    background: #1E2130;
    border: 1px solid #2D3748;
    border-radius: 12px;
    padding: 1rem !important;
}
[data-testid="stMetricValue"] {
    color: #4F8BF9 !important;
    font-weight: 700 !important;
}
</style>
""", unsafe_allow_html=True)


# ── Helper functions ──────────────────────────────────────────────────────────

def try_parse_number(value):
    if not isinstance(value, str):
        return value
    v = value.strip()
    if v in ("", "-"):
        return value
    negative = False
    if v.startswith("(") and v.endswith(")"):
        v = v[1:-1]; negative = True
    elif v.startswith("-"):
        v = v[1:]; negative = True
    v = re.sub(r"[Rp$€£¥\s]", "", v)
    if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)?$", v):
        v = v.replace(".", "").replace(",", ".")
    elif re.match(r"^\d{1,3}(,\d{3})*(\.\d+)?$", v):
        v = v.replace(",", "")
    elif re.match(r"^\d+(,\d+)$", v):
        v = v.replace(",", ".")
    elif not re.match(r"^\d+(\.\d+)?$", v):
        return value
    try:
        num = float(v)
        if negative: num = -num
        return int(num) if num == int(num) else num
    except ValueError:
        return value


def extract_tables(pdf_bytes):
    results = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total_pages = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if not tables:
                continue
            for tbl_idx, table in enumerate(tables, start=1):
                if not table or len(table) < 2:
                    continue
                header = [str(h).strip() if h else "" for h in table[0]]
                rows   = [[str(c).strip() if c else "" for c in row] for row in table[1:]]
                seen   = {}
                clean_header = []
                for col in header:
                    if col in seen:
                        seen[col] += 1
                        clean_header.append(f"{col}_{seen[col]}")
                    else:
                        seen[col] = 0
                        clean_header.append(col)
                rows = [r for r in rows if any(c.strip() for c in r)]
                parsed_rows = [[try_parse_number(c) for c in row] for row in rows]
                results.append({
                    "page": page_num,
                    "table_index": tbl_idx,
                    "headers": clean_header,
                    "rows": parsed_rows,
                    "total_pages": total_pages,
                })
    return results


def build_excel(tables):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        current_row = 0
        for item in tables:
            df = pd.DataFrame(item["rows"], columns=item["headers"])
            label_df = pd.DataFrame([[f"[ Halaman {item['page']} — Tabel {item['table_index']} ]"]])
            label_df.to_excel(writer, sheet_name="Laporan Keuangan",
                              startrow=current_row, index=False, header=False)
            current_row += 1
            df.to_excel(writer, sheet_name="Laporan Keuangan",
                        startrow=current_row, index=False, header=True)
            current_row += len(df) + 1 + 2
    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Laporan Keuangan"]
    _apply_styles(ws)
    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.read()


def _apply_styles(ws):
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_rows  = set()
    header_rows = set()
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and str(cell.value).startswith("[ Halaman"):
            label_rows.add(cell.row)
    for r in label_rows:
        header_rows.add(r + 1)
    data_row_counters = {}
    for row in ws.iter_rows():
        row_num   = row[0].row
        is_label  = row_num in label_rows
        is_header = row_num in header_rows
        if not is_label and not is_header:
            block_start = max((h for h in header_rows if h < row_num), default=0)
            data_row_counters.setdefault(block_start, 0)
            data_row_counters[block_start] += 1
            is_even = (data_row_counters[block_start] % 2 == 0)
        else:
            is_even = False
        for cell in row:
            if is_label:
                cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
                cell.fill      = PatternFill("solid", fgColor="2F5496")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif is_header:
                cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                cell.fill      = PatternFill("solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border
            else:
                cell.font = Font(size=10, name="Calibri")
                cell.fill = PatternFill("solid", fgColor="D9E1F2" if is_even else "FFFFFF")
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment     = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) and cell.value != int(cell.value) else '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len    = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
    for row in ws.row_dimensions.values():
        row.height = 18


# ── UI ────────────────────────────────────────────────────────────────────────

# Header: nama pembuat di kanan atas
try:
    from PIL import Image
    import base64
    img = Image.open("foto.jpg")
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    img_b64 = base64.b64encode(buf.getvalue()).decode()
    foto_html = f'<img src="data:image/jpeg;base64,{img_b64}" style="width:48px;height:48px;border-radius:50%;object-fit:cover;border:2px solid #4F8BF9;">'
except:
    foto_html = '<div style="width:48px;height:48px;border-radius:50%;background:linear-gradient(135deg,#4F8BF9,#A855F7);display:flex;align-items:center;justify-content:center;font-size:1.2rem;">👤</div>'

st.markdown(f"""
<div style="display:flex;justify-content:flex-end;align-items:center;gap:10px;margin-bottom:0.5rem;">
    <div style="text-align:right;">
        <div style="font-size:0.85rem;font-weight:700;color:#FAFAFA;letter-spacing:1px;">R1C4</div>
    </div>
    {foto_html}
</div>
""", unsafe_allow_html=True)

# Hero
st.markdown("""
<div class="hero">
    <div class="hero-badge">⚡ PDF Converter</div>
    <h1 class="hero-title">Ubah PDF jadi<br><span class="gradient-text">Excel dalam detik</span></h1>
    <p class="hero-sub">
        Upload laporan keuangan PDF — semua tabel diekstrak otomatis,<br>
        diformat rapi, dan siap pakai rumus Excel.
    </p>
</div>
""", unsafe_allow_html=True)

# Feature chips
st.markdown("""
<div class="features">
    <div class="feature-chip">✅ Format angka otomatis</div>
    <div class="feature-chip">🎨 Styling Excel rapi</div>
    <div class="feature-chip">📊 Multi-tabel & Multi-halaman</div>
    <div class="feature-chip">🔢 Siap rumus SUM/AVERAGE</div>
    <div class="feature-chip">🆓 100% Gratis</div>
</div>
<div class="glow-line"></div>
""", unsafe_allow_html=True)

# Upload
st.markdown('<p class="upload-label">📂 Upload File PDF</p>', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Drag & drop PDF di sini atau klik untuk browse",
    type=["pdf"],
    label_visibility="collapsed",
)

# Jika file diupload
if uploaded_file:
    file_size = uploaded_file.size / 1024 / 1024

    # Info file
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-icon">�</div>
            <span class="stat-value">PDF</span>
            <span class="stat-label">Format</span>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-icon">💾</div>
            <span class="stat-value">{file_size:.1f} MB</span>
            <span class="stat-label">Ukuran</span>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="stat-card">
            <div class="stat-icon">📎</div>
            <span class="stat-value" style="font-size:0.85rem">{uploaded_file.name[:16]}...</span>
            <span class="stat-label">File</span>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("⚡ Konversi ke Excel Sekarang", use_container_width=True, type="primary"):
        progress = st.progress(0, text="Membaca file PDF...")
        time.sleep(0.3)

        try:
            pdf_bytes = uploaded_file.read()
            progress.progress(20, text="Mengekstrak tabel...")

            tables = extract_tables(pdf_bytes)
            progress.progress(60, text="Memformat data...")

            if not tables:
                progress.empty()
                st.error("❌ Tidak ada tabel yang ditemukan di PDF ini. Pastikan PDF bukan hasil scan.")
            else:
                excel_bytes = build_excel(tables)
                progress.progress(90, text="Menyiapkan file Excel...")
                time.sleep(0.3)
                progress.progress(100, text="Selesai!")
                time.sleep(0.4)
                progress.empty()

                # Hitung total baris data
                total_rows = sum(len(t["rows"]) for t in tables)
                total_pages = tables[-1]["total_pages"] if tables else 0

                # Stats hasil
                st.markdown('<div class="result-card">', unsafe_allow_html=True)
                st.markdown('<div class="result-title">🎉 Konversi berhasil!</div>', unsafe_allow_html=True)

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.metric("📊 Tabel", len(tables))
                with c2:
                    st.metric("📄 Halaman", total_pages)
                with c3:
                    st.metric("📝 Baris Data", total_rows)

                st.markdown("<br>", unsafe_allow_html=True)

                output_name = uploaded_file.name.replace(".pdf", ".xlsx")
                st.download_button(
                    label="📥 Download Excel Sekarang",
                    data=excel_bytes,
                    file_name=output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)

        except Exception as e:
            progress.empty()
            st.error(f"❌ Terjadi kesalahan: {str(e)}")

# Divider & footer
st.markdown("""
<div class="divider"></div>
<div class="footer">
    © 2026 r1c4enterprise · All Rights Reserved
</div>
""", unsafe_allow_html=True)
